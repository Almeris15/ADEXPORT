import logging
import sys
import os
import pandas as pd
from datetime import datetime
from collections import defaultdict
from log_config import setup_logging
from openpyxl.styles import PatternFill, Border, Side, Alignment


# ----- Variables -----
dir = os.path.dirname(__file__)  # Emplacement du fichier grâce à son chemin absolue
dirExcel = dir + "\\..\\ExportExcel\\"
dirTemp = dir + "\\..\\ImportCSV\\"
OutputFile = sys.argv[1]  # Nom du fichier Excel = 1er argument en appel de script
color = 'F2F2F2'  # Color for the cell in hex code
colorborder = 'FF000000'  # color of the border of cells
rotation = 45  # Witch angle to rotate the text in a cell
widths = 4.3  # widths of all columns
widthsFirst = 20  # widths of the first column
# ----- Fin des variables -----


# ----- Début des fonctions -----
def Call_files():
    """Fonction pour appel de fichier."""
    file_names_list = []
    file_groups_list = []
    file_membership_list = []
    for file_name in os.listdir(dirTemp):  # se place dans le dossier Temp
        if file_name.startswith("ADmembers_") and file_name.endswith(".csv"):  # ne choisit que les fichiers csv commençant par Admembers_
            if file_name.split('_')[1] == OutputFile:  # Vérifie que le ADNameDNS donné en argument de programme soit dans le nom du fichier
                file_names_list.append(file_name)
        if file_name.startswith("ADgroups_") and file_name.endswith(".csv"):  # ne choisit que les fichiers csv commençant par Adroups_
            if file_name.split('_')[1] == OutputFile:  # Vérifie que le ADNameDNS donné en argument de programme soit dans le nom du fichier
                file_groups_list.append(file_name)
        if file_name.startswith("ADmembership_") and file_name.endswith(".csv"):  # ne choisit que les fichiers csv commençant par Admembership_
            if file_name.split('_')[1] == OutputFile:  # Vérifie que le ADNameDNS donné en argument de programme soit dans le nom du fichier
                file_membership_list.append(file_name)
    file_ADnames_list_sort = sorted(file_names_list,
                                    key=lambda x: datetime.strptime(x.split('_')[-1].split(".")[0], '%Y%m%d'),
                                    reverse=True)  # Sort la liste de fichier trié par date : du plus récent au plus ancien
    file_ADgroups_list_sort = sorted(file_groups_list,
                                    key=lambda x: datetime.strptime(x.split('_')[-1].split(".")[0], '%Y%m%d'),
                                    reverse=True)  # Sort la liste de fichier trié par date : du plus récent au plus ancien
    file_ADmembership_list_sort = sorted(file_membership_list,
                                        key=lambda x: datetime.strptime(x.split('_')[-1].split(".")[0], '%Y%m%d'),
                                        reverse=True)  # Sort la liste de fichier trié par date : du plus récent au plus ancien
    return file_ADnames_list_sort, file_ADgroups_list_sort, file_ADmembership_list_sort  # return les 3 fichier csv nécessaires


def search_file_list(file_list):
    """Fonction pour avoir les 2 fichier CSV de chaque list pour comparaison."""
    file2 = file_list[0]
    file1 = file_list[1]
    print(file1, file2)  # print chaque fichier comparé pour la console
    logging.info('Compare %s et %s', file1, file2) # print dans le fichier de log
    file_list.pop(0)
    file_date_str = file2.split("_")[-1].split(".")[0]  # Présume que le format est "ADmembers_'Ad.Name'_YYYYMMDD.csv" et garde que la date
    file_date = int(file_date_str)
    return file1, file2, file_date


def FileToList(filecsv):
    """Fonction pour les 2 premier fichier csv avec les User et les Groupes."""
    file_read = os.path.join(dirTemp, filecsv)  # Path absolue de l'emplacement des fichier csv
    with open(file_read, 'r') as fileR:
        List = fileR.readlines()
    ListFinal = [i.replace('"', '').replace("\n", '') for i in List[1:]]  # Structure le code en enlevant les quotes et \n en enlève la 1ère ligne
    return ListFinal


def concat_lists(listA, listB):
    """Assemble les 2 listes."""
    listR = listA[:]  # prend la valeur listA
    seen = set(listR)
    listBOnly = [x for x in listB if x not in seen and not seen.add(x)]  # si dans listB mais pas dans la listA
    listAOnly = [x for x in listA if x not in listB]  # si dans listA mais pas dans la listB
    listR.extend(listBOnly)  # ajout de listBOnly a liste A pour avoir toute la liste avec 1 seule ocuurence
    listR.sort()  # trie de la liste
    return listR, listBOnly, listAOnly


def Group_Membership(filecsv):
    """Fonction pour le 3ème fichier csv avec groupemembership."""
    file_read = os.path.join(dirTemp, filecsv)  # Path absolue de l'emplacement des fichier csv
    with open(file_read, 'r') as fileR:
        lineR = fileR.readlines()
        dico = defaultdict(list)  # création d'un dict par défaut, c'est un type de dict
        for lineR in map(str.rstrip, lineR):
            Name, Group = lineR.split(",")  # Sépare la liste csv en 2 colonnes à ","
            Name = Name.strip('"')  # Enlève les quotes
            Group = Group.strip('"')  # Enlève les quotes
            dico[Group].append(Name)
    del dico['GroupName']
    return dict(dico)


def Generate_table(list1, list2, dico):
    """Fonction pour générer un tableau grâce aux csv."""
    tableau = []
    for name in list2:
        row = [name in dico.get(group, []) and 'X' or ' ' for group in list1]  # Met des X si user appartient au group sinon met rien / espace
        tableau.append(row)  # Crée la ligne
    table = pd.DataFrame(tableau, columns=list1, index=list2)  # Crée la matrice
    return table


def bordure_cell(cell, rightborder, leftborder, topborder, bottomborder, colors):
    """Set border of cell."""
    cell.border = Border(right=Side(border_style=rightborder, color=colors),
                        left=Side(border_style=leftborder, color=colors),
                        top=Side(border_style=topborder, color=colors),
                        bottom=Side(border_style=bottomborder, color=colors))


def Compare_table(df1, df2, OutputExcel):
    """Fonction pour comparer les 2 dataframes."""
    mask1 = df1 == "X"  # prend la position de tout les X dans la matrice 1
    mask2 = df2 == "X"  # prend la position de tout les X dans la matrice 2

    writer = pd.ExcelWriter(OutputExcel, engine="openpyxl")  # création d'un fichier excel de sortie
    df2.to_excel(writer, index=True)  # mettre la matrice la plus récente dans le excel
    ws = writer.book.active

    # --- début des variables de compare_table ---
    red_fill = PatternFill(start_color="FF0000",  # setup couleur rouge
                            end_color="FF0000",
                            fill_type="solid")
    green_fill = PatternFill(start_color="00FF00",  # setup couleur verte
                             end_color="00FF00",
                             fill_type="solid")
    list_column_green = []
    list_column_red = []
    list_row_green = []
    list_row_red = []
    # --- Fin des variables de compare_table ---


    # --- début des fonctions de compare ---
    def fill_column(list_column, fill):
        """Fonction pour remplir les colonnes."""
        if list_column != []:  # si la liste n'est pas vide
            for col_num in list_column:
                # Prend chaque cellule de chaque ligne correspondant au numéro de colonne
                for row in ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
                    for cell in row:
                        cell.fill = fill  # colore la cellule


    def fill_row(list_row, fill):
        """Fonction pour remplir les lignes."""
        if list_row != []:  # si la liste n'est pas vide
            for row_num in list_row:
                for cell in ws[row_num][1:]:  # prend chaque cellule de la ligne
                    cell.fill = fill  # colore la cellule


    def fill_column_with_color(cell_axe, listcolor, listcoloraxe, color):
        """Fonction pour coloré les colonnes défini à remplir."""
        for cell in ws[1]:
            if cell.value in listcolor:  # si la valeur est dans la liste
                cell.fill = color  # Set color
                listcoloraxe.append(cell_axe)  # ajoute dans la liste


    def fill_row_with_color(cell_axe, listcolor, listcoloraxe, color):
        """Fonction pour colore les lignes défini à remplir."""
        for cell in ws['A']:
            if cell.value in listcolor:  # si la valeur est dans la liste
                cell.fill = color  # Set color
                listcoloraxe.append(cell_axe)  # ajoute dans la liste
    # --- fin des fonctions de compare ---


    for row_index, row in enumerate(mask1.values):  # prend mask 1 avec la position de tout les X non déposé dans le excel
        for col_index, cell in enumerate(row):
            current_cell = ws.cell(row=row_index + 2, column=col_index + 2)
            current_cell.alignment = Alignment(horizontal='center', vertical='center')  # centre le texte
            bordure_cell(current_cell, None, None, 'thin', 'thin', colorborder)
            if row_index % 2 != 0:
                # mise en page du excel pour toutes les lignes paires
                current_cell.fill = PatternFill(patternType='solid', fgColor=color)  # Set color
            if cell and not mask2.iloc[row_index, col_index]:
                # compare la position des X entre les 2 matrices, condition si dans la 1ère matrice et non dans la 2ème
                current_cell.fill = red_fill  # Set color
            elif not cell and mask2.iloc[row_index, col_index]:
                # compare la position des X entre les 2 matrices, condition si dans la 2ème matrice et non dans la 1ère
                current_cell.fill = green_fill  # Set color

    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = widths  # set la largeur de chaque colonnes
    ws.column_dimensions['A'].width = widthsFirst  # set la largeur de la 1ère colonne

    for cell in ws[1]:  # Prend la 1ère ligne
        cell.alignment = Alignment(textRotation=rotation)  # rotation of the text for the group name
        bordure_cell(cell, 'thin', 'thin', None, 'thin', colorborder)  # Set border

    for cell in ws['A']:  # Prend la 1ère colonne
        cell.alignment = Alignment(horizontal='left')
        bordure_cell(cell, 'thin', None, 'thin', 'thin', colorborder)  # Set border

    fill_column_with_color(cell.column, listUserGreen, list_column_green, green_fill)
    fill_column_with_color(cell.column, listUserRed, list_column_red, red_fill)
    fill_row_with_color(cell.row, listUserGreen, list_row_green, green_fill)
    fill_row_with_color(cell.row, listUserRed, list_row_red, red_fill)
    fill_column(list_column_green, green_fill)
    fill_column(list_column_red, red_fill)
    fill_row(list_row_green, green_fill)
    fill_row(list_row_red, red_fill)

    writer.close()
# ----- Fin des fonctions -----


# ----- Appels aux fonctions -----
setup_logging()

User_file_list, Groups_file_list, GroupMembership_file_list = Call_files()

while len(User_file_list) > 1:  # Quand il y a + qu'un fichier dans la liste de fichier alors comparaison
    # retourne les fichiers csv nécessaires pour la comparaison
    User_file1, User_file2, user_file_date = search_file_list(User_file_list)
    Group_file1, Group_file2, group_file_date = search_file_list(Groups_file_list)
    Membership_file1, Membership_file2, membership_file_date = search_file_list(GroupMembership_file_list)

    if user_file_date == group_file_date == membership_file_date:  # Condition pour s'assurer que les fichier ont la même date
        ListUser1 = FileToList(User_file1)
        ListUser2 = FileToList(User_file2)
        ListGroup1 = FileToList(Group_file1)
        ListGroup2 = FileToList(Group_file2)
        dict1 = Group_Membership(Membership_file1)
        dict2 = Group_Membership(Membership_file2)

        ListUserFinal, listUserGreen, listUserRed = concat_lists(ListUser1, ListUser2)
        ListGroupFinal, listGroupGreen, listGroupRed = concat_lists(ListGroup1, ListGroup2)

        table1 = Generate_table(ListGroupFinal, ListUserFinal, dict1)
        table2 = Generate_table(ListGroupFinal, ListUserFinal, dict2)

        date = str(user_file_date)
        fileName = OutputFile + "_" + date + ".xlsx"
        Excelfile = os.path.join(dirExcel, fileName)

        tableFinal = Compare_table(table1, table2, Excelfile)
        print(f'Create "{fileName}"')  # print the file edit
        logging.info('Created file %s',fileName)
    else:
        print("!! Erreur au niveau des dates des fichiers csv !!")  # print si il y a une erreur
        logging.error("!! Error date file CSV !!")  # log si il y a une erreur


print("Compare.py Finish")  # print end of program
logging.info('End Compare.py')  # log la fin du programme
