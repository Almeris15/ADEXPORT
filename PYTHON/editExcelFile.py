import os
import sys
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Border, Side, Alignment


# --- Début des Variables ---
dir = os.path.dirname(__file__)  # Emplacement du fichier dans l'ordinateur/serveur
dirExcel = dir + "\\..\\ExportExcel\\"  # Emplacement de la racine du programme
InputFile = sys.argv[1] + "_" + datetime.now().strftime("%Y%m%d") + ".xlsx"
ExcelFile = os.path.join(dirExcel, InputFile)  # Name of the Excel file and his place to this file
color = 'F2F2F2'  # Color for the cell in hex code
colorborder = 'FF000000'  # color of the border of cells
rotation = 45  # Witch angle to rotate the text in a cell
widths = 4.3  # widths of all columns
widthsFirst = 20  # widths of the first column
# --- Fin de Variables ---


# --- Début des fonctions ---
# --- Fin des fonctions ---


# Open the Excel file
wb = openpyxl.load_workbook(ExcelFile)
ws = wb.active  # Select the active sheet

# Iterate through all cells in the sheet
for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center the text in the cell

# Iterate through all rows and columns in the sheet
for i, row in enumerate(ws.rows):
    for j, cell in enumerate(row):
        if i % 2 != 0:  # If the row and column are even
            cell.fill = PatternFill(patternType='solid', fgColor=color)  # Color the cell / change hex to modify the color
            cell.border = Border(top=Side(border_style='thin', color=colorborder),  # Set border
                                bottom=Side(border_style='thin', color=colorborder))
        else:
            cell.fill = PatternFill(patternType=None, fgColor=color)  # Color the cell / change hex to modify the color)

for column in ws.columns:
    ws.column_dimensions[column[0].column_letter].width = widths  # Set the width of all columns to x width (x is not pixel)

ws.column_dimensions['A'].width = widthsFirst  # Set the width of the First column to x width (x is not pixel)

# First Row
for cell in ws[1]:
    cell.alignment = Alignment(textRotation=rotation)  # Set the text rotation to x degrees
    cell.border = Border(right=Side(border_style='thin', color=colorborder),  # Set border
                        left=Side(border_style='thin', color=colorborder),
                        bottom=Side(border_style='thin', color=colorborder))
# First Column
for cell in ws['A']:
    cell.alignment = Alignment(horizontal='left')  # Center the text in the cell
    cell.border = Border(right=Side(border_style='thin', color=colorborder),  # Set border
                        top=Side(border_style='thin', color=colorborder),
                        bottom=Side(border_style='thin', color=colorborder))

wb.save(ExcelFile)  # Save and close the changes to the Excel file
